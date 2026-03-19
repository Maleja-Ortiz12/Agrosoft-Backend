from sqlalchemy.orm import Session
from typing import Optional
from models import models
from schemas import schemas

def get_animal(db: Session, animal_id: int):
    animal = db.query(models.Animal).filter(models.Animal.id == animal_id).first()
    if animal:
        if animal.mother:
            animal.mother_breed = animal.mother.breed
        
        # Obtener última inseminación para estado de gestación
        latest_ins = db.query(models.Insemination).filter(models.Insemination.animal_id == animal.id).order_by(models.Insemination.date.desc()).first()
        if latest_ins and latest_ins.status == "Confirmada (Cargada)":
            animal.pregnancy_status = "Cargada"
            animal.estimated_birth_date = latest_ins.estimated_birth_date
        else:
            animal.pregnancy_status = "Sin cargar"
            animal.estimated_birth_date = None
            
    return animal

def get_animal_by_tag(db: Session, tag_id: str):
    return db.query(models.Animal).filter(models.Animal.tag_id == tag_id).first()

def get_animals(db: Session, skip: int = 0, limit: int = 100, include_inactive: bool = False):
    query = db.query(models.Animal)
    if not include_inactive:
        query = query.filter(models.Animal.status == "Activo")
    animals = query.offset(skip).limit(limit).all()
    for a in animals:
        if a.mother:
            a.mother_breed = a.mother.breed
        
        # Obtener última inseminación para cada animal
        latest_ins = db.query(models.Insemination).filter(models.Insemination.animal_id == a.id).order_by(models.Insemination.date.desc()).first()
        if latest_ins and latest_ins.status == "Confirmada (Cargada)":
            a.pregnancy_status = "Cargada"
            a.estimated_birth_date = latest_ins.estimated_birth_date
        else:
            a.pregnancy_status = "Sin cargar"
            a.estimated_birth_date = None
            
    return animals

def create_animal(db: Session, animal: schemas.AnimalCreate):
    db_animal = models.Animal(**animal.dict())
    db.add(db_animal)
    db.commit()
    db.refresh(db_animal)
    return db_animal

def update_animal(db: Session, animal_id: int, update: schemas.AnimalUpdate):
    db_animal = db.query(models.Animal).filter(models.Animal.id == animal_id).first()
    if db_animal:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_animal, key, value)
        db.commit()
        db.refresh(db_animal)
    return db_animal

def delete_animal(db: Session, animal_id: int):
    db_animal = db.query(models.Animal).filter(models.Animal.id == animal_id).first()
    if db_animal:
        db.delete(db_animal)
        db.commit()
    return db_animal

def get_sales(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Sale).offset(skip).limit(limit).all()

def create_sale(db: Session, sale: schemas.SaleCreate):
    db_sale = models.Sale(**sale.dict())
    db.add(db_sale)
    db.commit()
    db.refresh(db_sale)
    return db_sale

def update_sale(db: Session, sale_id: int, update: schemas.SaleUpdate):
    db_sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if db_sale:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_sale, key, value)
        db.commit()
        db.refresh(db_sale)
    return db_sale

def delete_sale(db: Session, sale_id: int):
    db_sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if db_sale:
        db.delete(db_sale)
        db.commit()
    return db_sale

def get_milk_prediction(db: Session):
    # Por ahora, una predicción simple basada en el promedio de los últimos 7 registros de leche
    recent_sales = db.query(models.Sale).filter(models.Sale.type == "Leche").order_by(models.Sale.date.desc()).limit(7).all()
    
    if not recent_sales:
        return 0.0
    
    total_quantity = sum(s.quantity for s in recent_sales)
    return round(total_quantity / len(recent_sales), 2)

def get_vaccinations(db: Session, animal_id: Optional[int] = None):
    query = db.query(models.Vaccination)
    if animal_id:
        query = query.filter(models.Vaccination.animal_id == animal_id)
    return query.all()

def create_vaccination(db: Session, vaccination: schemas.VaccinationCreate):
    db_vac = models.Vaccination(**vaccination.dict())
    db.add(db_vac)
    db.commit()
    db.refresh(db_vac)
    return db_vac

def update_vaccination(db: Session, vaccination_id: int, update: schemas.VaccinationUpdate):
    db_vac = db.query(models.Vaccination).filter(models.Vaccination.id == vaccination_id).first()
    if db_vac:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_vac, key, value)
        db.commit()
        db.refresh(db_vac)
    return db_vac

def delete_vaccination(db: Session, vaccination_id: int):
    db_vaccination = db.query(models.Vaccination).filter(models.Vaccination.id == vaccination_id).first()
    if db_vaccination:
        db.delete(db_vaccination)
        db.commit()
    return db_vaccination

def create_bulk_vaccinations(db: Session, bulk_vaccine: schemas.BulkVaccinationCreate):
    created_records = []
    for animal_id in bulk_vaccine.animal_ids:
        db_vac = models.Vaccination(
            animal_id=animal_id,
            record_type=bulk_vaccine.record_type,
            treatment_name=bulk_vaccine.treatment_name,
            date=bulk_vaccine.date,
            cost=bulk_vaccine.cost,
            observation=bulk_vaccine.observation,
            next_due_date=bulk_vaccine.next_due_date
        )
        db.add(db_vac)
        created_records.append(db_vac)
    db.commit()
    for record in created_records:
        db.refresh(record)
    return created_records

def get_health_alerts(db: Session, days: int = 7):
    from datetime import date, timedelta
    today = date.today()
    limit_date = today + timedelta(days=days)
    
    alerts = db.query(models.Vaccination).filter(
        models.Vaccination.next_due_date <= limit_date
    ).order_by(models.Vaccination.next_due_date.asc()).all()
    
    result = []
    for a in alerts:
        a_dict = {
            "id": a.id,
            "animal_id": a.animal_id,
            "record_type": a.record_type,
            "treatment_name": a.treatment_name,
            "next_due_date": str(a.next_due_date),
            "animal_name": a.animal.name if a.animal else "Sin nombre",
            "animal_tag": a.animal.tag_id if a.animal else "N/A"
        }
        result.append(a_dict)
    
    return result

def get_fertilization_alerts(db: Session, days: int = 7):
    from datetime import date, timedelta
    today = date.today()
    limit_date = today + timedelta(days=days)
    
    alerts = db.query(models.Fertilization).filter(
        models.Fertilization.next_due_date <= limit_date
    ).order_by(models.Fertilization.next_due_date.asc()).all()
    
    result = []
    for a in alerts:
        a_dict = {
            "id": a.id,
            "crop_id": a.crop_id,
            "crop_name": a.crop.crop_name if a.crop else "N/A",
            "fertilizer_name": a.fertilizer_name,
            "next_due_date": str(a.next_due_date),
            "products_json": a.products_json
        }
        result.append(a_dict)
    
    return result

def get_expenses(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Expense).offset(skip).limit(limit).all()

def create_expense(db: Session, expense: schemas.ExpenseCreate):
    db_expense = models.Expense(**expense.dict())
    db.add(db_expense)
    db.commit()
    db.refresh(db_expense)
    return db_expense

def update_expense(db: Session, expense_id: int, update: schemas.ExpenseUpdate):
    db_expense = db.query(models.Expense).filter(models.Expense.id == expense_id).first()
    if db_expense:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_expense, key, value)
        db.commit()
        db.refresh(db_expense)
    return db_expense

def delete_expense(db: Session, expense_id: int):
    db_expense = db.query(models.Expense).filter(models.Expense.id == expense_id).first()
    if db_expense:
        db.delete(db_expense)
        db.commit()
    return db_expense

def get_financial_summary(db: Session):
    # Ganadería
    lv_sales = db.query(models.Sale).filter(models.Sale.sector == "Ganadería").all()
    lv_total_sales = sum((s.quantity or 0.0) * (s.price or 0.0) for s in lv_sales)
    
    # Sumar Gastos base + Vacunas/Salud
    lv_base_expenses = db.query(models.Expense).filter(models.Expense.sector == "Ganadería").all()
    lv_vaccination_expenses = db.query(models.Vaccination).all()
    lv_total_expenses = sum(e.amount or 0.0 for e in lv_base_expenses) + sum(v.cost or 0.0 for v in lv_vaccination_expenses)
    
    # Agricultura
    # Sumar Ventas base + Cosechas
    ag_base_sales = db.query(models.Sale).filter(models.Sale.sector == "Agricultura").all()
    ag_harvest_sales = db.query(models.Harvest).all()
    ag_total_sales = sum((s.quantity or 0.0) * (s.price or 0.0) for s in ag_base_sales) + sum((h.quantity or 0.0) * (h.unit_cost or 0.0) for h in ag_harvest_sales)
    
    ag_expenses = db.query(models.Expense).filter(models.Expense.sector == "Agricultura").all()
    ag_total_expenses = sum(e.amount or 0.0 for e in ag_expenses)
    
    # Nómina - Sumar por sector del trabajador
    all_payroll = db.query(models.Payroll).all()
    lv_payroll_expenses = sum(p.amount or 0.0 for p in all_payroll if p.worker and p.worker.sector == "Ganadería")
    ag_payroll_expenses = sum(p.amount or 0.0 for p in all_payroll if p.worker and p.worker.sector == "Agricultura")

    return {
        "livestock": {
            "total_sales": round(lv_total_sales, 2),
            "total_expenses": round(lv_total_expenses + lv_payroll_expenses, 2),
            "balance": round(lv_total_sales - (lv_total_expenses + lv_payroll_expenses), 2)
        },
        "agriculture": {
            "total_sales": round(ag_total_sales, 2),
            "total_expenses": round(ag_total_expenses + ag_payroll_expenses, 2),
            "balance": round(ag_total_sales - (ag_total_expenses + ag_payroll_expenses), 2)
        },
        "total": {
            "total_sales": round(lv_total_sales + ag_total_sales, 2),
            "total_expenses": round(lv_total_expenses + ag_total_expenses + lv_payroll_expenses + ag_payroll_expenses, 2),
            "balance": round((lv_total_sales + ag_total_sales) - (lv_total_expenses + ag_total_expenses + lv_payroll_expenses + ag_payroll_expenses), 2)
        }
    }

def get_inseminations(db: Session, animal_id: Optional[int] = None):
    query = db.query(models.Insemination)
    if animal_id:
        query = query.filter(models.Insemination.animal_id == animal_id)
    results = query.all()
    for ins in results:
        if ins.cow:
            ins.animal_name = ins.cow.name
            ins.animal_breed = ins.cow.breed
    return results

from datetime import timedelta
def create_insemination(db: Session, insemination: schemas.InseminationCreate):
    # Cálculo gestación: 283 días promedio para vacas
    est_birth = insemination.date + timedelta(days=283)
    db_ins = models.Insemination(**insemination.dict())
    db_ins.estimated_birth_date = est_birth
    db.add(db_ins)
    db.commit()
    db.refresh(db_ins)
    return db_ins

def update_insemination(db: Session, insemination_id: int, update: schemas.InseminationUpdate):
    db_ins = db.query(models.Insemination).filter(models.Insemination.id == insemination_id).first()
    if db_ins:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_ins, key, value)
        db.commit()
        db.refresh(db_ins)
    return db_ins

def delete_insemination(db: Session, insemination_id: int):
    db_ins = db.query(models.Insemination).filter(models.Insemination.id == insemination_id).first()
    if db_ins:
        db.delete(db_ins)
        db.commit()
    return db_ins

def get_weight_records(db: Session, animal_id: Optional[int] = None):
    query = db.query(models.WeightRecord)
    if animal_id:
        query = query.filter(models.WeightRecord.animal_id == animal_id)
    return query.order_by(models.WeightRecord.date.desc()).all()

def create_weight_record(db: Session, weight_record: schemas.WeightRecordCreate):
    db_weight = models.WeightRecord(**weight_record.dict())
    db.add(db_weight)
    db.commit()
    db.refresh(db_weight)
    return db_weight

def update_weight_record(db: Session, weight_id: int, update: schemas.WeightRecordUpdate):
    db_weight = db.query(models.WeightRecord).filter(models.WeightRecord.id == weight_id).first()
    if db_weight:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_weight, key, value)
        db.commit()
        db.refresh(db_weight)
    return db_weight

def delete_weight_record(db: Session, weight_id: int):
    db_weight = db.query(models.WeightRecord).filter(models.WeightRecord.id == weight_id).first()
    if db_weight:
        db.delete(db_weight)
        db.commit()
    return db_weight

# --- CRUD AGRICULTURA ---

def get_crops(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Crop).offset(skip).limit(limit).all()

def create_crop(db: Session, crop: schemas.CropCreate):
    db_crop = models.Crop(**crop.dict())
    db.add(db_crop)
    db.commit()
    db.refresh(db_crop)
    return db_crop

def update_crop(db: Session, crop_id: int, update: schemas.CropUpdate):
    db_crop = db.query(models.Crop).filter(models.Crop.id == crop_id).first()
    if db_crop:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_crop, key, value)
        db.commit()
        db.refresh(db_crop)
    return db_crop

def delete_crop(db: Session, crop_id: int):
    db_crop = db.query(models.Crop).filter(models.Crop.id == crop_id).first()
    if db_crop:
        db.delete(db_crop)
        db.commit()
    return db_crop

def get_irrigations(db: Session, crop_id: Optional[int] = None):
    query = db.query(models.Irrigation)
    if crop_id:
        query = query.filter(models.Irrigation.crop_id == crop_id)
    return query.all()

def create_irrigation(db: Session, irrigation: schemas.IrrigationCreate):
    db_irrigation = models.Irrigation(**irrigation.dict())
    db.add(db_irrigation)
    db.commit()
    db.refresh(db_irrigation)
    return db_irrigation

def update_irrigation(db: Session, irrigation_id: int, update: schemas.IrrigationUpdate):
    db_irr = db.query(models.Irrigation).filter(models.Irrigation.id == irrigation_id).first()
    if db_irr:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_irr, key, value)
        db.commit()
        db.refresh(db_irr)
    return db_irr

def delete_irrigation(db: Session, irrigation_id: int):
    db_irr = db.query(models.Irrigation).filter(models.Irrigation.id == irrigation_id).first()
    if db_irr:
        db.delete(db_irr)
        db.commit()
    return db_irr

def get_fertilizations(db: Session, crop_id: Optional[int] = None):
    query = db.query(models.Fertilization)
    if crop_id:
        query = query.filter(models.Fertilization.crop_id == crop_id)
    return query.all()

def create_fertilization(db: Session, fertilization: schemas.FertilizationCreate):
    db_fert = models.Fertilization(**fertilization.dict())
    db.add(db_fert)
    db.commit()
    db.refresh(db_fert)
    return db_fert

def update_fertilization(db: Session, fertilization_id: int, update: schemas.FertilizationUpdate):
    db_fert = db.query(models.Fertilization).filter(models.Fertilization.id == fertilization_id).first()
    if db_fert:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_fert, key, value)
        db.commit()
        db.refresh(db_fert)
    return db_fert

def delete_fertilization(db: Session, fertilization_id: int):
    db_fert = db.query(models.Fertilization).filter(models.Fertilization.id == fertilization_id).first()
    if db_fert:
        db.delete(db_fert)
        db.commit()
    return db_fert

def get_pests(db: Session, crop_id: Optional[int] = None):
    query = db.query(models.Pest)
    if crop_id:
        query = query.filter(models.Pest.crop_id == crop_id)
    return query.all()

def create_pest(db: Session, pest: schemas.PestCreate):
    db_pest = models.Pest(**pest.dict())
    db.add(db_pest)
    db.commit()
    db.refresh(db_pest)
    return db_pest

def update_pest(db: Session, pest_id: int, update: schemas.PestUpdate):
    db_pest = db.query(models.Pest).filter(models.Pest.id == pest_id).first()
    if db_pest:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_pest, key, value)
        db.commit()
        db.refresh(db_pest)
    return db_pest

def delete_pest(db: Session, pest_id: int):
    db_pest = db.query(models.Pest).filter(models.Pest.id == pest_id).first()
    if db_pest:
        db.delete(db_pest)
        db.commit()
    return db_pest

def get_harvests(db: Session, crop_id: Optional[int] = None):
    query = db.query(models.Harvest)
    if crop_id:
        query = query.filter(models.Harvest.crop_id == crop_id)
    harvests = query.all()
    for h in harvests:
        if h.crop:
            h.crop_name = h.crop.crop_name
    return harvests

def create_harvest(db: Session, harvest: schemas.HarvestCreate):
    db_harvest = models.Harvest(**harvest.dict())
    db.add(db_harvest)
    db.commit()
    db.refresh(db_harvest)
    return db_harvest

def update_harvest(db: Session, harvest_id: int, update: schemas.HarvestUpdate):
    db_harvest = db.query(models.Harvest).filter(models.Harvest.id == harvest_id).first()
    if db_harvest:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_harvest, key, value)
        db.commit()
        db.refresh(db_harvest)
    return db_harvest

def delete_harvest(db: Session, harvest_id: int):
    db_harvest = db.query(models.Harvest).filter(models.Harvest.id == harvest_id).first()
    if db_harvest:
        db.delete(db_harvest)
        db.commit()
    return db_harvest

# --- CRUD PERSONAL ---

def get_workers(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Worker).offset(skip).limit(limit).all()

def create_worker(db: Session, worker: schemas.WorkerCreate):
    db_worker = models.Worker(**worker.dict())
    db.add(db_worker)
    db.commit()
    db.refresh(db_worker)
    return db_worker

def update_worker(db: Session, worker_id: int, update: schemas.WorkerUpdate):
    db_worker = db.query(models.Worker).filter(models.Worker.id == worker_id).first()
    if db_worker:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_worker, key, value)
        db.commit()
        db.refresh(db_worker)
    return db_worker

def delete_worker(db: Session, worker_id: int):
    db_worker = db.query(models.Worker).filter(models.Worker.id == worker_id).first()
    if db_worker:
        db.delete(db_worker)
        db.commit()
    return db_worker

def get_work_logs(db: Session, worker_id: Optional[int] = None, only_unpaid: bool = False):
    query = db.query(models.WorkLog)
    if worker_id:
        query = query.filter(models.WorkLog.worker_id == worker_id)
    if only_unpaid:
        query = query.filter(models.WorkLog.paid == False)
    return query.order_by(models.WorkLog.date.desc()).all()

def create_work_log(db: Session, work_log: schemas.WorkLogCreate):
    db_log = models.WorkLog(**work_log.dict())
    db.add(db_log)
    db.commit()
    db.refresh(db_log)
    return db_log

def create_work_logs_batch(db: Session, batch: schemas.WorkLogBatchCreate):
    db_logs = [
        models.WorkLog(
            worker_id=batch.worker_id,
            date=d,
            start_time=batch.start_time,
            end_time=batch.end_time,
            duties=batch.duties,
            calculated_cost=batch.calculated_cost
        )
        for d in batch.dates
    ]
    
    try:
        db.add_all(db_logs)
        db.commit()
        for log in db_logs:
            db.refresh(log)
        return db_logs
    except Exception as e:
        db.rollback()
        raise e

def update_work_log(db: Session, log_id: int, update: schemas.WorkLogUpdate):
    db_log = db.query(models.WorkLog).filter(models.WorkLog.id == log_id).first()
    if db_log:
        for key, value in update.dict(exclude_unset=True).items():
            setattr(db_log, key, value)
        db.commit()
        db.refresh(db_log)
    return db_log

def delete_work_log(db: Session, log_id: int):
    db_log = db.query(models.WorkLog).filter(models.WorkLog.id == log_id).first()
    if db_log:
        db.delete(db_log)
        db.commit()
    return db_log

import pandas as pd
import io

def get_tax_report_excel(db: Session):
    # Obtener todas las ventas
    sales = db.query(models.Sale).all()
    sales_data = [{
        "Fecha": s.date,
        "Tipo": "INGRESO",
        "Categoría": s.type,
        "Sector": s.sector,
        "Monto": (s.quantity or 0.0) * (s.price or 0.0),
        "Descripción": s.description or ""
    } for s in sales]

    # Obtener todos los gastos base
    expenses = db.query(models.Expense).all()
    expenses_data = [{
        "Fecha": e.date,
        "Tipo": "GASTO",
        "Categoría": e.category,
        "Sector": e.sector,
        "Monto": e.amount or 0.0,
        "Descripción": e.description or ""
    } for e in expenses]

    # Obtener gastos de vacunación
    vaccinations = db.query(models.Vaccination).filter(models.Vaccination.cost > 0).all()
    vaccination_data = [{
        "Fecha": v.date,
        "Tipo": "GASTO",
        "Categoría": "Salud/Vacunación",
        "Sector": "Ganadería",
        "Monto": v.cost or 0.0,
        "Descripción": f"{v.record_type}: {v.treatment_name} - {v.animal.name if v.animal else 'Animal'}"
    } for v in vaccinations]

    # Obtener gastos de nómina
    payrolls = db.query(models.Payroll).all()
    payroll_data = [{
        "Fecha": p.payment_date,
        "Tipo": "GASTO",
        "Categoría": "Nómina/Personal",
        "Sector": p.worker.sector if p.worker else "General",
        "Monto": p.amount or 0.0,
        "Descripción": f"Pago a {p.worker.name if p.worker else 'Trabajador'}"
    } for p in payrolls]

    # Obtener ingresos por cosechas
    harvests = db.query(models.Harvest).all()
    harvests_data = [{
        "Fecha": h.harvest_date,
        "Tipo": "INGRESO",
        "Categoría": "Cosecha",
        "Sector": "Agricultura",
        "Monto": (h.quantity or 0.0) * (h.unit_cost or 0.0),
        "Descripción": f"Cosecha de {h.crop.crop_name if h.crop else 'Cultivo'} ({h.quantity or 0.0} {h.unit or ''})"
    } for h in harvests]

    # Combinar todos los datos
    all_data = sales_data + harvests_data + expenses_data + vaccination_data + payroll_data
    if not all_data:
        return None

    df = pd.DataFrame(all_data)
    df = df.sort_values(by="Fecha")

    # Calcular Totales (Sincronizado con get_financial_summary)
    total_ingresos = sum(d['Monto'] for d in (sales_data + harvests_data))
    total_gastos = sum(d['Monto'] for d in (expenses_data + vaccination_data + payroll_data))
    balance_final = total_ingresos - total_gastos

    # Crear buffer de Excel
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte Tributario')
        
        # Obtener el workbook y el sheet para aplicar estilos
        workbook = writer.book
        sheet = workbook['Reporte Tributario']
        
        # Estilos
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ingreso_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde claro
        gasto_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rojo claro
        
        # Aplicar estilo al encabezado
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Aplicar colores a las filas según el tipo
        for row in range(2, len(df) + 2):
            tipo = sheet.cell(row=row, column=2).value # Columna 'Tipo'
            fill = ingreso_fill if tipo == "INGRESO" else gasto_fill
            for col in range(1, 7): # Columnas 1 a 6
                sheet.cell(row=row, column=col).fill = fill
        
        # Añadir Totales al final
        last_row = len(df) + 4
        sheet.cell(row=last_row, column=4).value = "TOTAL INGRESOS:"
        sheet.cell(row=last_row, column=4).font = Font(bold=True)
        sheet.cell(row=last_row, column=5).value = total_ingresos
        sheet.cell(row=last_row, column=5).font = Font(bold=True, color="006100")
        
        sheet.cell(row=last_row + 1, column=4).value = "TOTAL GASTOS:"
        sheet.cell(row=last_row + 1, column=4).font = Font(bold=True)
        sheet.cell(row=last_row + 1, column=5).value = total_gastos
        sheet.cell(row=last_row + 1, column=5).font = Font(bold=True, color="9C0006")
        
        sheet.cell(row=last_row + 2, column=4).value = "BALANCE FINAL:"
        sheet.cell(row=last_row + 2, column=4).font = Font(bold=True)
        sheet.cell(row=last_row + 2, column=5).value = balance_final
        sheet.cell(row=last_row + 2, column=5).font = Font(bold=True, size=12)
        
        # Ajustar ancho de columnas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    return output.getvalue()

# --- CRUD USUARIOS & AUTH ---
from passlib.context import CryptContext

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def get_user_by_username(db: Session, username: str):
    return db.query(models.User).filter(models.User.username == username).first()

def create_user(db: Session, user: schemas.UserCreate):
    hashed_password = pwd_context.hash(user.password)
    db_user = models.User(
        username=user.username,
        full_name=user.full_name,
        hashed_password=hashed_password
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

# --- CRUD NOMINA / PAYROLL ---
def create_payroll(db: Session, payroll: schemas.PayrollCreate):
    db_payroll = models.Payroll(
        worker_id=payroll.worker_id,
        amount=payroll.amount,
        payment_date=payroll.payment_date,
        signature_data=payroll.signature_data,
        log_ids=payroll.log_ids
    )
    db.add(db_payroll)
    
    # Mark associated logs as paid
    if payroll.log_ids:
        try:
            ids = [int(id_str.strip()) for id_str in payroll.log_ids.split(",") if id_str.strip().isdigit()]
            if ids:
                db.query(models.WorkLog).filter(models.WorkLog.id.in_(ids)).update({"paid": True}, synchronize_session=False)
        except Exception as e:
            print(f"Error marking logs as paid: {e}")

    db.commit()
    db.refresh(db_payroll)
    return db_payroll

def get_payroll_by_worker(db: Session, worker_id: int):
    return db.query(models.Payroll).filter(models.Payroll.worker_id == worker_id).order_by(models.Payroll.payment_date.desc()).all()

def delete_payroll(db: Session, payroll_id: int):
    db_payroll = db.query(models.Payroll).filter(models.Payroll.id == payroll_id).first()
    if db_payroll:
        db.delete(db_payroll)
        db.commit()
    return db_payroll
